import pytesseract
from PIL import Image
import re
import openpyxl
from datetime import date
import os

def process_screenshot(file_path):
    # Load the screenshot
    image = Image.open(file_path)

    # Extract text from the screenshot
    extracted_text = pytesseract.image_to_string(image)

    # Process the extracted text to get usernames and power levels
    user_data = re.findall(r"(\w+)\s+Power\s+(\d+)", extracted_text)
    
    return user_data

def update_spreadsheet(user_data_list, spreadsheet_name="user_data.xlsx"):
    # Load or create the spreadsheet
    if os.path.exists(spreadsheet_name):
        workbook = openpyxl.load_workbook(spreadsheet_name)
    else:
        workbook = openpyxl.Workbook()

    worksheet = workbook.active

    # Find the next available column
    column = 1
    while worksheet.cell(row=1, column=column).value:
        column += 1

    # Set the date as the column header
    worksheet.cell(row=1, column=column).value = date.today().strftime("%Y-%m-%d")

    # Add usernames and power levels to the spreadsheet
    row = 2
    for user_data in user_data_list:
        for username, power in user_data:
            worksheet.cell(row=row, column=column).value = f"{username}, {power}"
            row += 1

    # Save the spreadsheet
    workbook.save(spreadsheet_name)

# Set the path to the Tesseract executable
pytesseract.pytesseract.tesseract_cmd = "path/to/tesseract.exe"

# List the screenshot file paths
screenshots = ["screenshot1.png", "screenshot2.png", "screenshot3.png"]

# Process all screenshots
all_user_data = [process_screenshot(file_path) for file_path in screenshots]

# Update the spreadsheet with the new data
update_spreadsheet(all_user_data)
